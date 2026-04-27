"""Parse the FDA AI/ML-enabled device Excel file into structured records."""

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def parse_fda_excel(filepath: str) -> list[dict]:
    """Parse FDA AI/ML device list Excel file.

    The FDA publishes this at:
    https://www.fda.gov/medical-devices/software-medical-device-samd/
    artificial-intelligence-and-machine-learning-aiml-enabled-medical-devices

    Download the Excel file and place it in pipeline/data/
    """
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("Empty spreadsheet")
        return []

    # Find header row (first row with content)
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    # Map columns flexibly — FDA changes header names occasionally
    col_map = {}
    for i, h in enumerate(headers):
        if h == "submission number" or (h == "submission" and "submission_number" not in col_map):
            col_map["fda_submission_number"] = i
        elif h == "device" or ("device" in h and "name" in h):
            col_map["device_name"] = i
        elif "company" in h or "applicant" in h or "sponsor" in h:
            col_map["company"] = i
        elif "date" in h and ("decision" in h or "final" in h):
            col_map["decision_date"] = i
        elif "panel" in h or "specialty" in h:
            col_map["specialty_panel"] = i
        elif "product" in h and "code" in h:
            col_map["product_code"] = i

    # Prefer "Submission Number" (plain text) over "submission" (hyperlink formula)
    # The FDA file has both: col 1 is a hyperlink, col 2 is plain text
    for i, h in enumerate(headers):
        if h == "submission number":
            col_map["fda_submission_number"] = i
            break

    if "fda_submission_number" not in col_map:
        print("Warning: Could not map headers by name, using positional fallback")
        col_map = {
            "fda_submission_number": 2,
            "device_name": 3,
            "company": 4,
            "decision_date": 0,
            "specialty_panel": 5,
            "product_code": 6,
        }

    print(f"Column mapping: {col_map}")

    devices = []
    for row in rows[1:]:
        if not row or not row[col_map.get("fda_submission_number", 0)]:
            continue

        device = {}
        for field, col_idx in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                device[field] = str(val).strip()
            else:
                device[field] = None

        # Normalize date to ISO format
        if device.get("decision_date"):
            raw = device["decision_date"]
            if hasattr(raw, "isoformat"):
                device["decision_date"] = raw.isoformat()[:10]
            elif "/" in str(raw):
                # Handle MM/DD/YYYY format
                parts = str(raw).split("/")
                if len(parts) == 3:
                    device["decision_date"] = f"{parts[2]}-{parts[0].zfill(2)}-{parts[1].zfill(2)}"

        if device.get("fda_submission_number"):
            devices.append(device)

    wb.close()
    return devices


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python parse_fda_list.py <path_to_excel>")
        sys.exit(1)

    filepath = sys.argv[1]
    if not Path(filepath).exists():
        print(f"File not found: {filepath}")
        sys.exit(1)

    devices = parse_fda_excel(filepath)
    print(f"Parsed {len(devices)} devices")

    # Write to JSON for inspection
    out_path = Path(__file__).parent / "data" / "devices.json"
    out_path.parent.mkdir(exist_ok=True)
    with open(out_path, "w") as f:
        json.dump(devices, f, indent=2, default=str)
    print(f"Written to {out_path}")
